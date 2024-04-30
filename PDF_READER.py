import argparse
import PyPDF2
import openpyxl
import pandas as pd

def pdf_to_excel(pdf_path, excel_path, page_num=None):
    def read_pdf(pdf_path, page_num=None):
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            num_pages = len(pdf_reader.pages)

            if page_num is not None:
                if page_num < 1 or page_num > num_pages:
                    print(f"Error: Page number should be between 1 and {num_pages}")
                    return None
                else:
                    page = pdf_reader.pages[page_num - 1]
                    text = page.extract_text()
                    return text
            else:
                text = ""
                for page_num in range(num_pages):
                    page = pdf_reader.pages[page_num]
                    text += page.extract_text()
                return text

    def create_excel(excel_path):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.cell(row=1, column=1, value="Page Number")
        sheet.cell(row=1, column=2, value="Text")
        wb.save(excel_path)
        return wb, sheet

    def text_pdf_to_excel(text, excel_path):
        wb, sheet = create_excel(excel_path)
        text_list = text.split('\n')
        for row_num, line in enumerate(text_list, start=2):
            sheet.cell(row=row_num, column=1, value=row_num - 1)
            sheet.cell(row=row_num, column=2, value=line)
        wb.save(excel_path)
        
    def arrange_data_columns(excel_path):
        df = pd.read_excel(excel_path)
        selected_rows = df.iloc[[2, 3, 5, 17, 20, 23]]
        selected_rows = selected_rows.T
        selected_rows.columns = ['Invoice', 'Date', 'Insurance', "Patient's Name", 'Services', 'Balance']
        selected_rows.to_excel('output_file.xlsx', index=False)

    text = read_pdf(pdf_path, page_num)
    if text:
        text_pdf_to_excel(text, excel_path)
        arrange_data_columns(excel_path)
    else:
        print("No text extracted.")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Read contents of a PDF file.")
    parser.add_argument("file", help="Path to the PDF file")
    parser.add_argument("excel", help="Path to excel file")
    parser.add_argument("-r", "--read", action="store_true", help="Flag to read the PDF")
    parser.add_argument("-p", "--page", type=int, help="Page number to read (optional)")

    args = parser.parse_args()

    if args.read:
        pdf_to_excel(args.file, args.excel, args.page)
    else:
        print("Please provide the '-r' flag to read the PDF.")
